import sys
import json
import openpyxl

input_file  = sys.argv[1] if len(sys.argv) > 1 else "output.xlsx"
output_file = sys.argv[2] if len(sys.argv) > 2 else None

wb = openpyxl.load_workbook(input_file)
ws = wb.active

headers = [cell.value for cell in ws[1]]

representatives = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    rep = {}
    for key, val in zip(headers, row):
        if key == "electedYear":
            rep[key] = int(val) if val else ""
        else:
            rep[key] = val if val is not None else ""
    representatives.append(rep)

if output_file is None and representatives:
    year = representatives[0].get("electedYear", "")
    output_file = f"representatives_{year}.json" if year else "representatives.json"

with open(output_file, "w", encoding="utf-8") as f:
    json.dump(representatives, f, indent=2, ensure_ascii=False)

print(f"Saved {output_file} with {len(representatives)} representatives")
