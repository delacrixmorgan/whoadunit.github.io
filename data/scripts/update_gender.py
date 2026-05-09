"""
update_gender.py
----------------
Infers and updates the 'gender' column for every representative in both
the MP and ADUN sheets of a Malaysian representatives XLSX file.

Usage:
    python3 update_gender.py representative_2022.xlsx
"""

import re
import sys
import openpyxl

# ---------------------------------------------------------------------------
# Name-based gender rules
# ---------------------------------------------------------------------------

# Explicit Female names (full name, lower-cased, wiki markup stripped)
FEMALE_EXPLICIT = {
    # Malay female names
    "wan badariyah wan saad",
    "marzita mansor",
    "shamsilah siru",
    "halimaton shaadiah saad",
    "nor asilah mohamed zin",
    "siti ashah ghazali",
    "zuraida md noor",
    "rohani ibrahim",
    "hafidzah mustakim",
    "mardhiyyah johari",
    "nik asma' bahrum nik abdullah",
    "nor sham sulaiman",
    "salbiah mohamed",
    "norhaslinda zakaria",
    "wasanthee sinnasamy",
    "salina samsudin",
    "andansura rabu",
    "najihatussalehah ahmad",
    "bavani veraiah @ shasha",
    "angeline koo hai yen",
    "loh sze yee",
    "jenny choy tsi jen",
    "thulsi thivani manogaran",
    "wong may ing",
    "haslinda salleh",
    "ee chin li",
    "khairin nisa ismail",
    "nor rashidah ramli",
    "gan peck cheng",
    "sharifah azizah syed zain",
    "alwiyah talib",
    "rashidah ismail",
    "norlizah noh",
    "fauziah misri",
    "amira aisya abdul aziz",
    "liow cai tung",
    "marina ibrahim",
    "hasrunizah hassan",
    "julita majungki",
    "isnaraissah munirah majilis",
    "edna jessica majimbun",
    "dayang noorazah awang sohor",
    "sharifah hasidah sayeed aman ghazali",
    "violet yong wui wui",
    "rosey yunus",
    "rina jainal",
    "kalsom noordin",
    "tuminah kadi @ mohd hasim",
    "khaidiriah abu zahar",
    "siti faizah abdul azis",
    "ng kor sim",
    "saraswathy nallathanby",
    "fatimah abdullah",
    "maliaman kassim",
    "lisa hassan alban",
    # Chinese female names
    "gan ay ling",
    "heng lee lee",
    "lim siew khim",
    "connie tan hooi peng",
    "h'ng mooi lye",
    "phee syn tze",
    "wong yuee harng",
    "teh lai heng",
    "lee ting han",
    "chaw kam foon",
    "leng chau yen",
    "goh see hua",
    "lee kim shin",
    "sandrea ng shy ching",
    "wong chai yi",
}

# Names that look potentially female but are actually male
MALE_EXPLICIT = {
    "loi kok liang",
    "lok kok liang",
    "kumaran krishnan",
    "sundarajoo somu",
    "gooi zi sen",
    "gooi hsiao leung",
    "goh choon aik",
    "ho chi yang",
    "vidyananthan ramanadhan",
    "sivanesan achalingam",
    "jagdeep singh deo",
    "kumaresan aramugam",
    "raven kumar krishnasamy",
}

# Regex patterns (applied to lower-cased, cleaned name) indicating Female.
# Ordered from most specific to most general.
FEMALE_PATTERNS = [
    # Full-name patterns (must come first to avoid partial matches)
    r"\bnik asma\b",
    r"\bwan badariyah\b",
    r"\bsiti ashah\b",
    r"\bsiti faizah\b",
    r"\bng kor sim\b",
    r"\bheng lee lee\b",
    r"\bh'ng mooi\b",
    r"\bloh sze yee\b",
    r"\bjenny choy\b",
    r"\bphee syn tze\b",
    r"\blim siew khim\b",
    r"\bteh lai heng\b",
    r"\blee ting han\b",
    r"\bchaw kam foon\b",
    r"\bwong chai yi\b",
    r"\bleng chau yen\b",
    r"\bwong may\b",
    r"\bwong yuee\b",
    r"\bgan ay\b",
    r"\bgan peck\b",
    r"\bee chin\b",
    r"\bandansura\b",
    # Common first-name / title patterns
    r"\bsiti\b",
    r"\bnor\b",
    r"\bnoor\b",
    r"\bnorliza",
    r"\bnorhaslinda\b",
    r"\bnorazah\b",
    r"\bnoorazah\b",
    r"\bhafidzah\b",
    r"\brothani\b",
    r"\bfatimah\b",
    r"\bmarina\b",
    r"\bviolet\b",
    r"\bjenny\b",
    r"\bangeline\b",
    r"\bsandrea\b",
    r"\bmarzita\b",
    r"\bshamsilah\b",
    r"\bmardhiyyah\b",
    r"\bhalimaton\b",
    r"\bsalbiah\b",
    r"\bzuraida\b",
    r"\bsalina\b",
    r"\bnajihatussalehah\b",
    r"\bdayang\b",
    r"\brosey\b",
    r"\brina\b",
    r"\bjulita\b",
    r"\bisnaraissah\b",
    r"\bedna\b",
    r"\bhasrunizah\b",
    r"\bhaslinda\b",
    r"\bkalsom\b",
    r"\btuminah\b",
    r"\bkhaidiriah\b",
    r"\brashidah\b",
    r"\bnorlizah\b",
    r"\bfauziah\b",
    r"\balwiyah\b",
    r"\bsharifah\b",
    r"\bkhairin\b",
    r"\bamira\b",
    r"\bliow\b",
    r"\bthulsi\b",
    r"\bwasanthee\b",
    r"\bsaraswathy\b",
    r"\bbavani\b",
    r"\bmaliaman\b",
    r"\blisa\b",
    r"\bconnie\b",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_name(name: str) -> str:
    """Strip wiki markup like |link=...}} from a name string."""
    return re.sub(r"\|link=.*?\}\}", "", name or "").strip()


def assign_gender(name: str) -> str:
    """Return 'Female' or 'Male' based on name-pattern heuristics."""
    if not name:
        return "Male"

    cleaned = clean_name(name)
    lower = cleaned.lower()

    # Explicit male overrides take highest priority
    if lower in MALE_EXPLICIT:
        return "Male"

    # Explicit female list
    if lower in FEMALE_EXPLICIT:
        return "Female"

    # Regex pattern matching
    for pattern in FEMALE_PATTERNS:
        if re.search(pattern, lower):
            return "Female"

    # Default: Male (reflects majority of Malaysian elected representatives)
    return "Male"


def process_sheet(ws, sheet_label: str) -> tuple[int, int]:
    """
    Update the gender column (index 9, column J) for every data row in ws.
    Returns (female_count, male_count).
    """
    female_count = 0
    male_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name_cell = row[7]    # column H – name
        gender_cell = row[9]  # column J – gender

        gender = assign_gender(name_cell.value)
        gender_cell.value = gender

        if gender == "Female":
            female_count += 1
        else:
            male_count += 1

    print(f"  {sheet_label}: {female_count} Female, {male_count} Male  "
          f"(total {female_count + male_count})")
    return female_count, male_count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 update_gender.py <file.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    print(f"Loading '{filepath}' …")
    wb = openpyxl.load_workbook(filepath)

    expected_sheets = {"MP", "ADUN"}
    available = set(wb.sheetnames)
    sheets_to_process = expected_sheets & available

    if not sheets_to_process:
        print(f"ERROR: No MP or ADUN sheets found. Available sheets: {wb.sheetnames}")
        sys.exit(1)

    print("\nUpdating gender columns …")
    total_female = total_male = 0

    for sheet_name in sorted(sheets_to_process):   # MP before ADUN alphabetically reversed; sort gives ADUN, MP
        ws = wb[sheet_name]
        f, m = process_sheet(ws, sheet_name)
        total_female += f
        total_male += m

    wb.save(filepath)
    print(f"\nSaved '{filepath}'.")
    print(f"Grand total — Female: {total_female}, Male: {total_male}")

    # Summary: list all Female-assigned rows per sheet for quick review
    wb2 = openpyxl.load_workbook(filepath)
    for sheet_name in sorted(sheets_to_process):
        ws2 = wb2[sheet_name]
        females = [
            row[7] for row in ws2.iter_rows(min_row=2, values_only=True)
            if row[9] == "Female"
        ]
        print(f"\n--- {sheet_name}: Female representatives ({len(females)}) ---")
        for name in females:
            print(f"  {name}")


if __name__ == "__main__":
    main()
